import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "手順書"

# ページ設定（横向き）
ws.page_setup.orientation = 'landscape'
ws.page_setup.paperSize = ws.PAPERSIZE_A3

# 列幅設定 - 画面のレイアウトに合わせる
# A: 左端狭い列, B: 項目名ラベル列, C-H: サービス内容と手順（広い）, I-K: 留意事項
ws.column_dimensions['A'].width = 18   # 項目列
ws.column_dimensions['B'].width = 10   # 項目列の続き / 基本情報ラベル
ws.column_dimensions['C'].width = 14   # 基本情報値
ws.column_dimensions['D'].width = 14
ws.column_dimensions['E'].width = 14
ws.column_dimensions['F'].width = 8
ws.column_dimensions['G'].width = 6
ws.column_dimensions['H'].width = 10
ws.column_dimensions['I'].width = 16
ws.column_dimensions['J'].width = 10
ws.column_dimensions['K'].width = 16
ws.column_dimensions['L'].width = 30   # 留意事項

# スタイル定義
thin = Side(style='thin')
thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
header_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
dark_fill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')
header_font = Font(name='MS ゴシック', size=9, bold=True)
normal_font = Font(name='MS ゴシック', size=9)
title_font = Font(name='MS ゴシック', size=16, bold=True)
small_font = Font(name='MS ゴシック', size=8)
white_font = Font(name='MS ゴシック', size=9, bold=True, color='FFFFFF')

def apply_border(ws, row, col_start, col_end):
    for c in range(col_start, col_end + 1):
        ws.cell(row=row, column=c).border = thin_border

def make_label_cell(ws, row, col, value, merge_end_col=None):
    """ラベルセル（グレー背景）を作成"""
    cell = ws.cell(row=row, column=col)
    cell.value = value
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center')
    if merge_end_col:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_end_col)
        for c in range(col, merge_end_col + 1):
            ws.cell(row=row, column=c).border = thin_border
            ws.cell(row=row, column=c).fill = header_fill

def make_value_cell(ws, row, col, merge_end_col=None, value=''):
    """値セル（白背景）を作成"""
    cell = ws.cell(row=row, column=col)
    cell.value = value
    cell.font = normal_font
    cell.border = thin_border
    cell.alignment = Alignment(vertical='center')
    if merge_end_col:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_end_col)
        for c in range(col, merge_end_col + 1):
            ws.cell(row=row, column=c).border = thin_border

# ==========================================
# Row 1: 「定義様」(左) + 「手順書」(中央タイトル) + 「作成年月日 令和__年__月__日」(右)
# ==========================================
ws.row_dimensions[1].height = 30
ws.cell(row=1, column=1).value = '　　　　様'
ws.cell(row=1, column=1).font = Font(name='MS ゴシック', size=11)

ws.merge_cells('E1:H1')
ws.cell(row=1, column=5).value = '手順書'
ws.cell(row=1, column=5).font = title_font
ws.cell(row=1, column=5).alignment = Alignment(horizontal='center', vertical='center')

ws.merge_cells('J1:L1')
ws.cell(row=1, column=10).value = '作成年月日　令和　　年　　月　　日'
ws.cell(row=1, column=10).font = normal_font
ws.cell(row=1, column=10).alignment = Alignment(horizontal='right', vertical='center')

# ==========================================
# Row 3: 氏名 / 性別 / 生年月日 / 電話番号
# ==========================================
row = 3
ws.row_dimensions[row].height = 22

make_label_cell(ws, row, 2, '氏名')
make_value_cell(ws, row, 3, 5)  # B-E: 氏名値

make_label_cell(ws, row, 6, '性別')
make_value_cell(ws, row, 7)  # G: 性別値

make_label_cell(ws, row, 8, '生年月日')
make_value_cell(ws, row, 9, 10)  # I-J: 生年月日値

make_label_cell(ws, row, 11, '電話番号')
make_value_cell(ws, row, 12)  # L: 電話番号値

# ==========================================
# Row 4: 住所
# ==========================================
row = 4
ws.row_dimensions[row].height = 22

make_label_cell(ws, row, 2, '住所')
make_value_cell(ws, row, 3, 12)

# ==========================================
# Row 5: 実施期間
# ==========================================
row = 5
ws.row_dimensions[row].height = 22

make_label_cell(ws, row, 2, '実施期間')
make_value_cell(ws, row, 3, 12)

# ==========================================
# Row 7: テーブルヘッダー（項目 / サービス内容と手順 / 留意事項）
# ==========================================
row = 7
ws.row_dimensions[row].height = 22

# 項目ヘッダー (A-B)
make_label_cell(ws, row, 1, '項目', 2)

# サービス内容と手順ヘッダー (C-K)
make_label_cell(ws, row, 3, 'サービス内容と手順', 11)

# 留意事項ヘッダー (L)
make_label_cell(ws, row, 12, '留意事項')

# ==========================================
# 手順項目（空欄のテンプレート）
# ==========================================
items = [
    '入浴・送り入れ',
    '車椅子移乗',
    '居日デイの用意',
    '服薬',
    'シャンプー浴・足浴',
    '更衣布・保湿',
    '更衣',
    '食前調理・食事',
    '',
    '配膳',
    '口腔ケア・タオル拭き',
    '食器を洗う',
    '余暇時間',
    '用務',
    '21:30',
    '入眠準備・配薬',
    '21:30-7:00',
    '7:00',
    'トイレ・顔・手・指を暖かいタオルで拭く',
    '朝食',
    '口腔ケア・タオルふき・保湿',
    '配膳',
    '着替布・保湿',
    '更衣',
    'デイ送り出し',
]

start_row = 8
for i, item in enumerate(items):
    r = start_row + i
    ws.row_dimensions[r].height = 18

    # 項目列 (A-B merged)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    cell = ws.cell(row=r, column=1)
    cell.value = item
    cell.font = normal_font
    cell.alignment = Alignment(vertical='center')
    for c in range(1, 3):
        ws.cell(row=r, column=c).border = thin_border

    # サービス内容と手順 (C-K merged) - 空欄
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=11)
    for c in range(3, 12):
        ws.cell(row=r, column=c).border = thin_border
    ws.cell(row=r, column=3).font = normal_font
    ws.cell(row=r, column=3).alignment = Alignment(vertical='center', wrap_text=True)

    # 留意事項 (L) - 空欄
    ws.cell(row=r, column=12).border = thin_border
    ws.cell(row=r, column=12).font = normal_font
    ws.cell(row=r, column=12).alignment = Alignment(vertical='center', wrap_text=True)

# ==========================================
# 注意点セクション
# ==========================================
note_start = start_row + len(items) + 1
ws.cell(row=note_start, column=1).value = '注意点'
ws.cell(row=note_start, column=1).font = Font(name='MS ゴシック', size=9, bold=True, underline='single')

# 注意点記入欄（10行分の空行）
for i in range(1, 11):
    r = note_start + i
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
    ws.cell(row=r, column=1).font = small_font
    ws.cell(row=r, column=1).alignment = Alignment(vertical='center', wrap_text=True)
    ws.row_dimensions[r].height = 15

# ==========================================
# 印刷設定
# ==========================================
ws.print_area = f'A1:L{note_start + 10}'
ws.page_margins.left = 0.4
ws.page_margins.right = 0.4
ws.page_margins.top = 0.4
ws.page_margins.bottom = 0.4

# 保存
output_path = '/Users/koike/Desktop/シフト/手順書テンプレート.xlsx'
wb.save(output_path)
print(f'保存完了: {output_path}')
